import os
import io
import shutil
import tempfile

from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from rest_framework.test import APITestCase

from apps.accounts.models import User, UserRole
from apps.core.normalizers import normalize_parcel_code
from apps.data_imports.models import ImportJob, ImportRowAction, ImportRowResult, ImportStatus
from apps.data_imports.services.excel_importer import ExcelMasterImporter
from apps.parcels.models import Parcel
from apps.people.models import OwnershipType, ParcelOwnership, Person


class ImportAndNormalizerTests(TestCase):
    def test_normalize_parcel_code_variants(self):
        self.assertEqual(normalize_parcel_code('n19'), 'N-19')
        self.assertEqual(normalize_parcel_code(' N-019 '), 'N-19')
        self.assertEqual(normalize_parcel_code('C 40b'), 'C-40B')

    def test_import_dry_run(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Datos_Propietarios'
        ws.append(['PARCELA', 'NOMBRE COMPLETO', 'RUT', 'DV', 'TELEFONO', 'EMAIL'])
        ws.append(['B-01', 'JUAN PEREZ', '12345678', '5', '912345678', 'juan@example.com'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        try:
            wb.save(path)
            importer = ExcelMasterImporter(file_path=path, dry_run=True)
            job = importer.run()
        finally:
            if os.path.exists(path):
                os.remove(path)

        self.assertTrue(job.dry_run)
        self.assertGreaterEqual(job.total_inserted, 1)
        self.assertEqual(Parcel.objects.count(), 0)

    def test_import_ignores_inflated_empty_rows(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'ANOTACIONES'
        ws.append(['PARCELA', 'FECHA', 'ANOTACION/OBSERVACION'])
        ws.append(['B-01', '2026-07-01', 'Nota real'])
        ws.cell(row=1000, column=4).fill = PatternFill(fill_type='solid', fgColor='FFFFFF')

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        try:
            wb.save(path)
            importer = ExcelMasterImporter(file_path=path, dry_run=True, sheets=['ANOTACIONES'], empty_row_break_limit=25)
            job = importer.run()
        finally:
            if os.path.exists(path):
                os.remove(path)

        sheet_result = job.sheet_results.get(sheet_name='ANOTACIONES')
        structure = (job.details or {}).get('structure', {})
        check = next(item for item in structure.get('checks', []) if item['sheet_name'] == 'ANOTACIONES')
        self.assertEqual(sheet_result.rows_read, 1)
        self.assertEqual(check['row_count'], 1)
        self.assertGreater(check['excel_reported_row_count'], check['row_count'])

    def test_import_replaces_existing_primary_owner_for_parcel(self):
        parcel = Parcel.objects.create(codigo_parcela='B-01')
        old_person = Person.objects.create(nombre_completo='PROPIETARIO ANTERIOR')
        old_ownership = ParcelOwnership.objects.create(
            parcela=parcel,
            persona=old_person,
            tipo=OwnershipType.PRINCIPAL,
            is_active=True,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = 'Datos_Propietarios'
        ws.append(['PARCELA', 'NOMBRE COMPLETO', 'RUT', 'DV', 'TELEFONO', 'EMAIL'])
        ws.append(['B-01', 'PROPIETARIO NUEVO', '12345678', '', '', 'nuevo@example.com'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        try:
            wb.save(path)
            importer = ExcelMasterImporter(file_path=path, dry_run=False, sheets=['Datos_Propietarios'])
            job = importer.run()
        finally:
            if os.path.exists(path):
                os.remove(path)

        old_ownership.refresh_from_db()
        self.assertEqual(job.status, ImportStatus.SUCCESS)
        self.assertFalse(old_ownership.is_active)
        self.assertEqual(
            ParcelOwnership.objects.filter(parcela=parcel, tipo=OwnershipType.PRINCIPAL, is_active=True, is_deleted=False).count(),
            1,
        )
        self.assertTrue(
            ParcelOwnership.objects.filter(
                parcela=parcel,
                persona__nombre_completo='PROPIETARIO NUEVO',
                tipo=OwnershipType.PRINCIPAL,
                is_active=True,
                is_deleted=False,
            ).exists()
        )

    def test_weekly_maestro_import_auto_runs_preview_then_commit(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Datos_Propietarios'
        ws.append(['PARCELA', 'NOMBRE COMPLETO', 'RUT', 'DV', 'TELEFONO', 'EMAIL'])
        ws.append(['B-01', 'JUAN PEREZ', '12345678', '5', '912345678', 'juan@example.com'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        report_dir = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, report_dir, ignore_errors=True)
        out = io.StringIO()
        try:
            wb.save(path)
            call_command(
                'weekly_maestro_import',
                file=path,
                mode='auto',
                sheets='Datos_Propietarios',
                report_dir=report_dir,
                stdout=out,
            )
        finally:
            if os.path.exists(path):
                os.remove(path)

        self.assertTrue(Parcel.objects.filter(codigo_parcela_key='B-1').exists())
        self.assertEqual(ImportJob.objects.filter(source_file=os.path.basename(path)).count(), 2)
        self.assertTrue(os.listdir(report_dir))

    def test_weekly_maestro_import_auto_stops_when_preview_exceeds_errors(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Mora GC'
        ws.append(['PARCELA', 'MORA CG UF', 'TOTAL PESOS'])
        ws.append(['B-01', '1,5', 'no-es-numero'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        report_dir = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, report_dir, ignore_errors=True)
        try:
            wb.save(path)
            with self.assertRaises(CommandError):
                call_command(
                    'weekly_maestro_import',
                    file=path,
                    mode='auto',
                    sheets='Mora GC',
                    report_dir=report_dir,
                    stdout=io.StringIO(),
                )
        finally:
            if os.path.exists(path):
                os.remove(path)

        self.assertEqual(ImportJob.objects.filter(dry_run=False).count(), 0)
        self.assertTrue(os.listdir(report_dir))


class ImportApiFlowTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            email='operador@example.com',
            password='test123456',
            role=UserRole.OPERADOR,
        )
        self.client.force_authenticate(self.user)

    def _build_workbook_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Datos_Propietarios'
        ws.append(['PARCELA', 'NOMBRE COMPLETO', 'RUT', 'DV', 'TELEFONO', 'EMAIL'])
        ws.append(['B-01', 'JUAN PEREZ', '12345678', '5', '912345678', 'juan@example.com'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(path)
        return path

    def test_preview_then_run_upload_flow(self):
        file_path = self._build_workbook_file()
        try:
            with open(file_path, 'rb') as fh:
                preview_response = self.client.post(
                    '/api/v1/imports/jobs/preview-upload/',
                    data={'file': fh},
                    format='multipart',
                )

            self.assertEqual(preview_response.status_code, 201)
            upload_session_id = preview_response.data['upload_session']['id']
            self.assertTrue(upload_session_id)
            self.assertTrue(preview_response.data['preview_job']['dry_run'])

            run_response = self.client.post(
                '/api/v1/imports/jobs/run-upload/',
                data={'upload_session_id': upload_session_id},
                format='json',
            )
            self.assertEqual(run_response.status_code, 202)
            self.assertFalse(run_response.data['job']['dry_run'])
            self.assertEqual(run_response.data['job']['status'], ImportStatus.PENDING)

            call_command('process_import_jobs', max_jobs=1, max_seconds=30)
            job = ImportJob.objects.get(id=run_response.data['job']['id'])
            self.assertEqual(job.status, ImportStatus.SUCCESS)
            self.assertTrue(Parcel.objects.filter(codigo_parcela_key='B-1').exists())
        finally:
            if os.path.exists(file_path):
                os.remove(file_path)

    def test_skip_preview_upload_only_creates_upload_session(self):
        file_path = self._build_workbook_file()
        try:
            with open(file_path, 'rb') as fh:
                response = self.client.post(
                    '/api/v1/imports/jobs/preview-upload/',
                    data={'file': fh, 'skip_preview': 'true'},
                    format='multipart',
                )

            self.assertEqual(response.status_code, 201)
            self.assertTrue(response.data['upload_session']['id'])
            self.assertTrue(response.data['skipped_preview'])
            self.assertIsNone(response.data['preview_job'])
            self.assertEqual(ImportJob.objects.count(), 0)
        finally:
            if os.path.exists(file_path):
                os.remove(file_path)

    def test_preview_upload_uses_community_update_profile(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Datos_Propietarios'
        ws.append(['PARCELA', 'NOMBRE COMPLETO', 'RUT', 'DV', 'TELEFONO', 'EMAIL'])
        ws.append(['B-01', 'JUAN PEREZ', '12345678', '5', '912345678', 'juan@example.com'])
        works = wb.create_sheet('OBRAS')
        works.append(['PARCELA N', 'CORTAFUEGO', 'LIMPIEZA'])
        works.append(['B-99', 'SI', 'NO'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        try:
            wb.save(path)
            with open(path, 'rb') as fh:
                response = self.client.post(
                    '/api/v1/imports/jobs/preview-upload/',
                    data={'file': fh, 'profile': 'actualizacion_comunidad'},
                    format='multipart',
                )
        finally:
            if os.path.exists(path):
                os.remove(path)

        self.assertEqual(response.status_code, 201)
        selected_sheets = response.data['upload_session']['selected_sheets']
        self.assertCountEqual(
            selected_sheets,
            ['Mora GC', 'Datos_Propietarios', 'OTROS DUEÑOS', 'RESIDENTES', 'PPU_LOGOS', 'ANOTACIONES'],
        )
        processed_sheets = {item['sheet_name'] for item in response.data['preview_job']['sheet_results']}
        self.assertIn('Datos_Propietarios', processed_sheets)
        self.assertNotIn('OBRAS', processed_sheets)
        self.assertEqual(Parcel.objects.count(), 0)

    def test_preview_rejects_invalid_column_mapping_json(self):
        file_path = self._build_workbook_file()
        try:
            with open(file_path, 'rb') as fh:
                response = self.client.post(
                    '/api/v1/imports/jobs/preview-upload/',
                    data={'file': fh, 'column_mapping': '{invalid-json'},
                    format='multipart',
                )
            self.assertEqual(response.status_code, 400)
            self.assertIn('column_mapping', response.data.get('detail', ''))
        finally:
            if os.path.exists(file_path):
                os.remove(file_path)

    def test_preview_invalid_workbook_returns_failed_job_with_fatal_issue(self):
        upload = SimpleUploadedFile(
            'maestro.xlsx',
            b'not a valid workbook',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            '/api/v1/imports/jobs/preview-upload/',
            data={'file': upload},
            format='multipart',
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.data['preview_job']['status'], ImportStatus.FAILED)
        self.assertTrue(any(issue['severity'] == 'FATAL' for issue in response.data['preview_issues']))

    def test_import_records_row_result_for_invalid_required_decimal(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Mora GC'
        ws.append(['PARCELA', 'MORA CG UF', 'TOTAL PESOS'])
        ws.append(['B-01', '1,5', 'no-es-numero'])

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        try:
            wb.save(path)
            importer = ExcelMasterImporter(file_path=path, dry_run=False, sheets=['Mora GC'])
            job = importer.run()
        finally:
            if os.path.exists(path):
                os.remove(path)

        self.assertEqual(job.status, ImportStatus.FAILED)
        self.assertGreaterEqual(job.total_errors, 1)
        self.assertTrue(
            ImportRowResult.objects.filter(import_job=job, action=ImportRowAction.ERROR, row_number=2).exists()
        )

    def test_cancel_running_job(self):
        job = ImportJob.objects.create(
            source_file='test.xlsx',
            source_hash='hash',
            source_path='/tmp/test.xlsx',
            dry_run=False,
            status=ImportStatus.RUNNING,
            initiated_by=self.user,
        )

        response = self.client.post(f'/api/v1/imports/jobs/{job.id}/cancel/', data={}, format='json')
        self.assertEqual(response.status_code, 202)

        job.refresh_from_db()
        self.assertEqual(job.status, ImportStatus.CANCELLED)
        self.assertTrue((job.details or {}).get('cancel_requested'))

    def test_stop_and_terminate_aliases(self):
        for endpoint in ('stop', 'terminate', 'cancel_requested'):
            job = ImportJob.objects.create(
                source_file=f'{endpoint}.xlsx',
                source_hash='hash',
                source_path=f'/tmp/{endpoint}.xlsx',
                dry_run=False,
                status=ImportStatus.PENDING,
                initiated_by=self.user,
            )
            response = self.client.post(f'/api/v1/imports/jobs/{job.id}/{endpoint}/', data={}, format='json')
            self.assertEqual(response.status_code, 202)
            job.refresh_from_db()
            self.assertEqual(job.status, ImportStatus.CANCELLED)

    def test_cancel_terminal_job_returns_conflict(self):
        job = ImportJob.objects.create(
            source_file='done.xlsx',
            source_hash='hash',
            source_path='/tmp/done.xlsx',
            dry_run=False,
            status=ImportStatus.SUCCESS,
            initiated_by=self.user,
        )

        response = self.client.post(f'/api/v1/imports/jobs/{job.id}/cancel/', data={}, format='json')
        self.assertEqual(response.status_code, 409)

    def test_post_detail_endpoint_is_backward_compatible(self):
        job = ImportJob.objects.create(
            source_file='legacy.xlsx',
            source_hash='hash',
            source_path='/tmp/legacy.xlsx',
            dry_run=True,
            status=ImportStatus.RUNNING,
            initiated_by=self.user,
        )

        response = self.client.post(f'/api/v1/imports/jobs/{job.id}/', data={}, format='json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(str(response.data['id']), str(job.id))
