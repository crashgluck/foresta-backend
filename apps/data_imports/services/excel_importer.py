from __future__ import annotations

import hashlib
import logging
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import unicodedata
from zipfile import BadZipFile

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import IntegrityError, transaction
from django.db.models import Count, Sum
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.datetime import from_excel

from apps.core.normalizers import normalize_email, normalize_parcel_code, normalize_phone, normalize_rut_dv, normalize_rut_number
from apps.core.validators import validate_rut
from apps.data_imports.models import ImportIssue, ImportJob, ImportRowAction, ImportRowResult, ImportSheetResult, ImportStatus, IssueSeverity
from apps.finance.models import CommonExpenseDebt, PaymentAgreement, ServiceDebt, ServiceType, UnpaidFine
from apps.notes.models import AdministrativeNote, NoteType
from apps.parcels.models import Parcel
from apps.people.models import OwnershipType, ParcelOwnership, ParcelResident, Person, ResidentType
from apps.utilities.models import CutType, ServiceCut, ServiceHistory
from apps.vehicles.models import Vehicle
from apps.works.models import ParcelWorkStatus

logger = logging.getLogger(__name__)


class ImportCancelledError(Exception):
    pass


@dataclass
class Counter:
    rows_read: int = 0
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    warnings: int = 0


GLOBAL_COLUMN_ALIASES = {
    'parcela': ['parcela', 'parcela n', 'n parcela', 'numero parcela', 'codigo parcela', 'lote', 'codigo lote', 'cliente'],
    'parcela n': ['parcela n', 'parcela', 'n parcela', 'numero parcela', 'codigo parcela'],
    'cliente': ['cliente', 'parcela', 'codigo cliente', 'codigo parcela'],
    'nombre completo': ['nombre completo', 'nombre propietario', 'propietario', 'dueno', 'dueño', 'nombre', 'razon social'],
    'rut': ['rut', 'run', 'rut propietario', 'rut persona'],
    'dv': ['dv', 'digito', 'digito verificador'],
    'telefono fijo': ['telefono fijo', 'telefono', 'fono', 'tel fijo'],
    'telefono movil': ['telefono movil', 'celular', 'movil', 'telefono celular'],
    'e mail': ['e mail', 'email', 'correo', 'correo electronico', 'mail'],
    'email': ['email', 'e mail', 'correo', 'correo electronico', 'mail'],
    'residente': ['residente', 'estado residente', 'tipo residente', 'habitante'],
    'observaciones': ['observaciones', 'obs', 'comentarios', 'comentario'],
    'ppu': ['ppu', 'patente', 'placa', 'placa patente'],
    'marca': ['marca', 'marca vehiculo'],
    'tipo': ['tipo', 'tipo vehiculo', 'clase'],
    'color': ['color', 'color vehiculo'],
    'codigo': ['codigo', 'codigo acceso', 'tag', 'rfid'],
    'mora cg uf': ['mora cg uf', 'mora gc uf', 'mora ggcc uf', 'mora uf'],
    'total pesos': ['total pesos', 'total peso', 'total clp', 'total $', 'total'],
    'total deuda': ['total deuda', 'saldo total', 'deuda total'],
    'total mora': ['total mora', 'mora total', 'saldo mora'],
    'saldo monto': ['saldo monto', 'saldo', 'monto saldo', 'monto'],
    'fecha': ['fecha', 'fecha evento', 'fecha registro'],
    'anotacion': ['anotacion', 'anotación', 'nota', 'comentario', 'texto'],
    'descripcion': ['descripcion', 'descripción', 'detalle', 'glosa'],
    'solicitante': ['solicitante', 'solicita', 'cliente solicitante'],
    'cortafuego': ['cortafuego', 'corta fuego'],
    'limpieza': ['limpieza', 'limpio'],
}


class ExcelMasterImporter:
    SHEET_REQUIREMENTS = {
        'Datos_Propietarios': ['parcela', 'nombre completo', 'rut'],
        'OTROS DUEÑOS': [],
        'RESIDENTES': ['parcela', 'residente'],
        'PPU_LOGOS': ['parcela', 'ppu'],
        'Mora GC': ['parcela', 'mora cg uf', 'total pesos'],
        'DESUDAS AyS': ['parcela', 'total deuda'],
        'MORA CONVENIO': ['parcela', 'total mora'],
        'Multas-Convenios impagas': ['parcela', 'empresa', 'saldo monto'],
        'Cortes Vigentes': ['cliente', 'estado'],
        'HISTORICO AYS': ['parcela', 'solicitante', 'descripcion'],
        'ANOTACIONES': ['parcela', 'fecha', 'anotacion'],
        'OBRAS': ['parcela n', 'cortafuego', 'limpieza'],
    }

    def __init__(
        self,
        file_path: str,
        dry_run: bool = False,
        initiated_by=None,
        sheets: list[str] | None = None,
        column_mapping: dict | None = None,
        log_success_rows: bool | None = None,
        empty_row_break_limit: int | None = None,
    ):
        self.file_path = Path(file_path)
        self.dry_run = dry_run
        self.initiated_by = initiated_by
        self.sheets_filter = {s.strip() for s in sheets} if sheets else None
        self.column_mapping = self._normalize_column_mapping(column_mapping or {})
        self.log_success_rows = getattr(settings, 'IMPORT_LOG_SUCCESS_ROWS', False) if log_success_rows is None else log_success_rows
        self.empty_row_break_limit = empty_row_break_limit or getattr(settings, 'IMPORT_EMPTY_ROW_BREAK_LIMIT', 150)
        self._cancel_check_every = 25
        self._operations_since_cancel_check = 0
        self._seen_keys: dict[str, dict[str, int]] = {}
        self._row_action_counts: dict[str, int] = {}

    def _parser_map(self):
        return {
            'Datos_Propietarios': self._parse_datos_propietarios,
            'OTROS DUEÑOS': self._parse_otros_duenos,
            'RESIDENTES': self._parse_residentes,
            'PPU_LOGOS': self._parse_vehiculos,
            'Mora GC': self._parse_mora_gc,
            'DESUDAS AyS': self._parse_deudas_ays,
            'MORA CONVENIO': self._parse_mora_convenio,
            'Multas-Convenios impagas': self._parse_multas,
            'Cortes Vigentes': self._parse_cortes,
            'HISTORICO AYS': self._parse_historico_ays,
            'ANOTACIONES': self._parse_anotaciones,
            'OBRAS': self._parse_obras,
        }

    def inspect_structure(self, workbook=None):
        should_close_workbook = workbook is None
        workbook = workbook or load_workbook(self.file_path, data_only=True, read_only=True)
        try:
            parser_map = self._parser_map()
            checks = []
            selected_unknown = sorted((self.sheets_filter or set()) - set(parser_map.keys()))
            for sheet_name in parser_map.keys():
                if self.sheets_filter and sheet_name not in self.sheets_filter:
                    continue
                required_keywords = self.SHEET_REQUIREMENTS.get(sheet_name, [])
                if sheet_name not in workbook.sheetnames:
                    checks.append(
                        {
                            'sheet_name': sheet_name,
                            'exists': False,
                            'header_found': False,
                            'required_keywords': required_keywords,
                            'missing_keywords': required_keywords,
                            'header_row': None,
                        }
                    )
                    continue
                ws = workbook[sheet_name]
                header_row, headers = self._find_header(ws, required_keywords) if required_keywords else (1, {})
                missing = []
                if required_keywords and headers:
                    missing = [
                        keyword
                        for keyword in required_keywords
                        if not any(self._header_matches(key, self._aliases_for(keyword, sheet_name)) for key in headers.keys())
                    ]
                elif required_keywords:
                    missing = list(required_keywords)

                checks.append(
                    {
                        'sheet_name': sheet_name,
                        'exists': True,
                        'header_found': bool(header_row),
                        'required_keywords': required_keywords,
                        'missing_keywords': missing,
                        'header_row': header_row or None,
                        'row_count': self._count_data_rows(ws, (header_row or 1) + 1, max_col=max(headers.values()) if headers else ws.max_column),
                        'excel_reported_row_count': max(ws.max_row - (header_row or 1), 0),
                        'columns': list(headers.keys()),
                    }
                )

            processable = [check for check in checks if check.get('exists') and check.get('header_found') and not check.get('missing_keywords')]
            return {
                'available_sheets': workbook.sheetnames,
                'selected_unknown_sheets': selected_unknown,
                'checks': checks,
                'processable_sheets': [check['sheet_name'] for check in processable],
                'is_structurally_valid': not selected_unknown and bool(processable),
            }
        finally:
            if should_close_workbook:
                workbook.close()

    def run(self, job: ImportJob | None = None) -> ImportJob:
        if not self.file_path.exists():
            raise FileNotFoundError(f'Archivo no encontrado: {self.file_path}')

        source_hash = self._hash_file(self.file_path)
        self._row_action_counts = {}
        logger.info(
            'Iniciando importacion de maestro Excel: file=%s dry_run=%s user=%s',
            self.file_path.name,
            self.dry_run,
            getattr(self.initiated_by, 'id', None),
        )
        if job is None:
            job = ImportJob.objects.create(
                source_file=self.file_path.name,
                source_hash=source_hash,
                source_path=str(self.file_path),
                dry_run=self.dry_run,
                status=ImportStatus.RUNNING,
                initiated_by=self.initiated_by,
            )
        else:
            job.sheet_results.all().delete()
            job.issues.all().delete()
            job.row_results.all().delete()
            details = dict(job.details or {})
            details.pop('summary', None)
            details.pop('fatal_errors', None)
            job.source_file = job.source_file or self.file_path.name
            job.source_hash = job.source_hash or source_hash
            job.source_path = job.source_path or str(self.file_path)
            job.dry_run = self.dry_run
            job.status = ImportStatus.RUNNING
            job.finished_at = None
            job.total_inserted = 0
            job.total_updated = 0
            job.total_skipped = 0
            job.total_errors = 0
            job.total_warnings = 0
            job.details = details
            if self.initiated_by and not job.initiated_by_id:
                job.initiated_by = self.initiated_by
            job.save(
                update_fields=[
                    'source_file',
                    'source_hash',
                    'source_path',
                    'dry_run',
                    'status',
                    'finished_at',
                    'total_inserted',
                    'total_updated',
                    'total_skipped',
                    'total_errors',
                    'total_warnings',
                    'details',
                    'initiated_by',
                ]
            )

        parser_map = self._parser_map()

        try:
            workbook = load_workbook(self.file_path, data_only=True, read_only=True)
        except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError) as exc:
            logger.warning('Archivo Excel invalido: %s (%s)', self.file_path, exc)
            return self._fail_job_with_fatal(job, 'invalid_workbook', f'No se pudo abrir el Excel: {exc}')

        structure = self.inspect_structure(workbook=workbook)
        details = dict(job.details or {})
        details.update(
            {
                'import_mode': 'preview' if self.dry_run else 'commit',
                'selected_sheets': sorted(self.sheets_filter) if self.sheets_filter else [],
                'structure': structure,
            }
        )
        job.details = details
        job.save(update_fields=['details'])

        if not self._validate_structure_for_run(job, structure):
            self._finalize_job(job)
            logger.info('Importacion abortada por errores estructurales: job=%s', job.id)
            return job

        cancelled = False
        for sheet_name, parser in parser_map.items():
            if self.sheets_filter and sheet_name not in self.sheets_filter:
                continue
            if self._is_cancel_requested(job):
                cancelled = True
                break
            if sheet_name not in workbook.sheetnames:
                self._issue(job, None, IssueSeverity.WARNING, sheet_name, None, None, 'sheet_missing', 'Hoja no encontrada')
                continue

            sheet_result = ImportSheetResult.objects.create(import_job=job, sheet_name=sheet_name, status=ImportStatus.RUNNING)
            counter = Counter()
            ws = workbook[sheet_name]

            try:
                parser(ws, job, sheet_result, counter)
                sheet_result.status = ImportStatus.PARTIAL if counter.errors else ImportStatus.SUCCESS
            except ImportCancelledError:
                cancelled = True
                sheet_result.status = ImportStatus.CANCELLED
                self._issue(
                    job,
                    sheet_result,
                    IssueSeverity.WARNING,
                    sheet_name,
                    None,
                    None,
                    'job_cancelled',
                    'Importación detenida por solicitud de cancelación.',
                )
            except Exception as exc:  # pragma: no cover
                logger.exception('Error importando hoja %s', sheet_name)
                counter.errors += 1
                self._issue(
                    job,
                    sheet_result,
                    IssueSeverity.ERROR,
                    sheet_name,
                    None,
                    None,
                    'sheet_crash',
                    f'Error crítico en hoja: {exc}',
                )
                sheet_result.status = ImportStatus.FAILED

            sheet_result.rows_read = counter.rows_read
            sheet_result.inserted = counter.inserted
            sheet_result.updated = counter.updated
            sheet_result.skipped = counter.skipped
            sheet_result.errors = counter.errors
            sheet_result.warnings = counter.warnings
            sheet_result.summary = (
                f'rows={counter.rows_read}, inserted={counter.inserted}, updated={counter.updated}, '
                f'skipped={counter.skipped}, errors={counter.errors}, warnings={counter.warnings}'
            )
            sheet_result.save()
            if cancelled:
                break

        try:
            self._finalize_job(job, cancelled=cancelled)
            logger.info(
                'Importacion finalizada: job=%s status=%s rows=%s created=%s updated=%s skipped=%s errors=%s warnings=%s',
                job.id,
                job.status,
                (job.details or {}).get('summary', {}).get('total_rows_read'),
                job.total_inserted,
                job.total_updated,
                job.total_skipped,
                job.total_errors,
                job.total_warnings,
            )
            return job
        finally:
            workbook.close()

    def _fail_job_with_fatal(self, job: ImportJob, error_code: str, message: str) -> ImportJob:
        self._issue(job, None, IssueSeverity.FATAL, 'WORKBOOK', None, None, error_code, message)
        self._row_result(job, None, None, ImportRowAction.ERROR, message, entity='WORKBOOK', issue_codes=[error_code])
        job.status = ImportStatus.FAILED
        job.finished_at = timezone.now()
        details = dict(job.details or {})
        details['fatal_errors'] = [{'code': error_code, 'message': message}]
        job.details = details
        job.total_errors = 1
        job.save(update_fields=['status', 'finished_at', 'details', 'total_errors'])
        return job

    def _validate_structure_for_run(self, job: ImportJob, structure: dict) -> bool:
        fatal_errors = []
        if not structure.get('available_sheets'):
            fatal_errors.append(('empty_workbook', 'El archivo no contiene hojas legibles.'))

        for sheet_name in structure.get('selected_unknown_sheets', []):
            fatal_errors.append(('unknown_selected_sheet', f'La hoja seleccionada "{sheet_name}" no esta soportada por el importador.'))

        if not structure.get('processable_sheets'):
            fatal_errors.append(('no_processable_sheets', 'No hay hojas procesables: faltan hojas soportadas o encabezados obligatorios.'))

        if fatal_errors:
            for code, message in fatal_errors:
                self._issue(job, None, IssueSeverity.FATAL, 'WORKBOOK', None, None, code, message)
                self._row_result(job, None, None, ImportRowAction.ERROR, message, entity='WORKBOOK', issue_codes=[code])
            details = dict(job.details or {})
            details['fatal_errors'] = [{'code': code, 'message': message} for code, message in fatal_errors]
            job.details = details
            job.save(update_fields=['details'])
            return False
        return True

    def _normalize_column_mapping(self, value):
        normalized = {}
        if not isinstance(value, dict):
            return normalized
        for sheet_name, aliases in value.items():
            sheet_key = self._norm_header(sheet_name)
            if not isinstance(aliases, dict):
                continue
            normalized[sheet_key] = {}
            for source_alias, mapped_alias in aliases.items():
                source_key = self._norm_header(source_alias)
                if not source_key:
                    continue
                if isinstance(mapped_alias, list):
                    normalized[sheet_key][source_key] = [self._norm_header(item) for item in mapped_alias if self._norm_header(item)]
                else:
                    mapped_key = self._norm_header(mapped_alias)
                    if mapped_key:
                        normalized[sheet_key][source_key] = [mapped_key]
        return normalized

    def _finalize_job(self, job: ImportJob, cancelled: bool = False):
        aggregates = job.sheet_results.aggregate(
            rows_read=Sum('rows_read'),
            inserted=Sum('inserted'),
            updated=Sum('updated'),
            skipped=Sum('skipped'),
            errors=Sum('errors'),
            warnings=Sum('warnings'),
        )
        issue_counts = {row['severity']: row['total'] for row in job.issues.values('severity').annotate(total=Count('id'))}
        job.total_inserted = aggregates['inserted'] or 0
        job.total_updated = aggregates['updated'] or 0
        job.total_skipped = aggregates['skipped'] or 0
        job.total_errors = max(aggregates['errors'] or 0, issue_counts.get(IssueSeverity.ERROR, 0) + issue_counts.get(IssueSeverity.FATAL, 0))
        job.total_warnings = max(aggregates['warnings'] or 0, issue_counts.get(IssueSeverity.WARNING, 0))
        job.finished_at = timezone.now()
        cancelled = cancelled or self._is_cancel_requested(job)
        action_counts = dict(self._row_action_counts)
        if not action_counts:
            action_counts = {row['action']: row['total'] for row in job.row_results.values('action').annotate(total=Count('id'))}
        rows_read = aggregates['rows_read'] or 0
        error_rows = job.row_results.filter(action=ImportRowAction.ERROR, row_number__isnull=False).values('sheet_name', 'row_number').distinct().count()
        details = dict(job.details or {})
        details['summary'] = {
            'total_rows_read': rows_read,
            'total_valid': max(rows_read - error_rows, 0),
            'total_imported': job.total_inserted + job.total_updated,
            'total_new': job.total_inserted,
            'total_updated': job.total_updated,
            'total_skipped': job.total_skipped,
            'total_errors': job.total_errors,
            'total_warnings': job.total_warnings,
            'fatal_errors': issue_counts.get(IssueSeverity.FATAL, 0),
            'row_actions': action_counts,
        }
        job.details = details

        if cancelled:
            job.status = ImportStatus.CANCELLED
            details['cancel_requested'] = True
            details.setdefault('cancelled_at', timezone.now().isoformat())
            job.details = details
        elif job.total_errors == 0:
            job.status = ImportStatus.SUCCESS
        elif job.total_inserted > 0 or job.total_updated > 0:
            job.status = ImportStatus.PARTIAL
        else:
            job.status = ImportStatus.FAILED

        job.save(
            update_fields=[
                'total_inserted',
                'total_updated',
                'total_skipped',
                'total_errors',
                'total_warnings',
                'finished_at',
                'status',
                'details',
            ]
        )

    def _is_cancel_requested(self, job: ImportJob) -> bool:
        job.refresh_from_db(fields=['status', 'details'])
        details = job.details or {}
        return job.status == ImportStatus.CANCELLED or bool(details.get('cancel_requested'))

    def _raise_if_cancel_requested(self, job: ImportJob, force: bool = False):
        self._operations_since_cancel_check += 1
        if not force and self._operations_since_cancel_check < self._cancel_check_every:
            return
        self._operations_since_cancel_check = 0
        if self._is_cancel_requested(job):
            raise ImportCancelledError()

    def _hash_file(self, path: Path) -> str:
        digest = hashlib.sha256()
        with path.open('rb') as fh:
            for chunk in iter(lambda: fh.read(4096), b''):
                digest.update(chunk)
        return digest.hexdigest()

    def _issue(
        self,
        job: ImportJob,
        sheet_result: ImportSheetResult | None,
        severity: str,
        sheet_name: str,
        row_number: int | None,
        column_name: str | None,
        error_code: str,
        message: str,
        raw_value: str = '',
    ):
        ImportIssue.objects.create(
            import_job=job,
            sheet_result=sheet_result,
            severity=severity,
            sheet_name=sheet_name,
            row_number=row_number,
            column_name=column_name or '',
            error_code=error_code,
            message=message,
            raw_value=(raw_value or '')[:500],
        )

    def _row_result(
        self,
        job: ImportJob,
        sheet_result: ImportSheetResult | None,
        row_number: int | None,
        action: str,
        message: str,
        *,
        entity: str = '',
        identifier: str = '',
        fields_affected: list | None = None,
        issue_codes: list | None = None,
    ):
        self._row_action_counts[action] = self._row_action_counts.get(action, 0) + 1
        if action not in {ImportRowAction.ERROR, ImportRowAction.WARNING} and not self.log_success_rows:
            return

        ImportRowResult.objects.create(
            import_job=job,
            sheet_result=sheet_result,
            sheet_name=sheet_result.sheet_name if sheet_result else 'WORKBOOK',
            row_number=row_number,
            entity=entity or '',
            record_identifier=str(identifier or '')[:255],
            action=action,
            message=message,
            fields_affected=fields_affected or [],
            issue_codes=issue_codes or [],
        )

    def _field_diff(self, instance, defaults: dict) -> list[dict]:
        changes = []
        for field, value in defaults.items():
            if value in (None, ''):
                continue
            old_value = getattr(instance, field, None)
            if old_value != value:
                changes.append({'field': field, 'before': str(old_value), 'after': str(value)})
        return changes

    def _norm_header(self, value) -> str:
        txt = '' if value is None else str(value)
        txt = unicodedata.normalize('NFKD', txt)
        txt = ''.join(ch for ch in txt if not unicodedata.combining(ch))
        txt = txt.replace('\n', ' ').replace('\r', ' ')
        txt = ''.join(ch if ch.isalnum() else ' ' for ch in txt)
        txt = ' '.join(txt.lower().split())
        return txt

    def _aliases_for(self, alias: str, sheet_name: str | None = None) -> list[str]:
        norm_alias = self._norm_header(alias)
        values = [norm_alias]
        values.extend(self._norm_header(item) for item in GLOBAL_COLUMN_ALIASES.get(norm_alias, []))
        if sheet_name:
            sheet_mapping = self.column_mapping.get(self._norm_header(sheet_name), {})
            values.extend(sheet_mapping.get(norm_alias, []))
        return [item for item in dict.fromkeys(values) if item]

    def _header_matches(self, header_key: str, aliases: list[str]) -> bool:
        return any(candidate == header_key or candidate in header_key or header_key in candidate for candidate in aliases)

    def _find_header(self, ws, required_keywords: list[str], max_rows: int = 20) -> tuple[int, dict[str, int]]:
        required_aliases = [self._aliases_for(keyword, ws.title) for keyword in required_keywords]
        for row_idx in range(1, max_rows + 1):
            values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
            header_map = {}
            for idx, val in enumerate(values, start=1):
                norm = self._norm_header(val)
                if norm:
                    header_map[norm] = idx
            if all(any(self._header_matches(key, aliases) for key in header_map.keys()) for aliases in required_aliases):
                return row_idx, header_map
        return 0, {}

    def _iter_data_row_numbers(self, ws, start_row: int, max_col: int | None = None):
        blank_streak = 0
        effective_max_col = max_col or ws.max_column
        for row_number, values in enumerate(
            ws.iter_rows(min_row=start_row, max_col=effective_max_col, values_only=True),
            start=start_row,
        ):
            if any(value not in (None, '') for value in values):
                blank_streak = 0
                yield row_number
                continue

            blank_streak += 1
            if blank_streak >= self.empty_row_break_limit:
                break

    def _count_data_rows(self, ws, start_row: int, max_col: int | None = None) -> int:
        return sum(1 for _ in self._iter_data_row_numbers(ws, start_row, max_col=max_col))

    def _cell(self, ws, row: int, col_map: dict[str, int], *aliases: str):
        for alias in aliases:
            lookup_aliases = self._aliases_for(alias, ws.title)
            for key, idx in col_map.items():
                if self._header_matches(key, lookup_aliases):
                    return ws.cell(row=row, column=idx).value
        return None

    def _to_int(
        self,
        value,
        default=0,
        *,
        job: ImportJob | None = None,
        sheet_result: ImportSheetResult | None = None,
        counter: Counter | None = None,
        row_number: int | None = None,
        column_name: str = '',
        required: bool = False,
    ):
        if value in (None, ''):
            if required and job and sheet_result and counter:
                counter.errors += 1
                self._issue(job, sheet_result, IssueSeverity.ERROR, sheet_result.sheet_name, row_number, column_name, 'missing_required_value', f'Valor requerido ausente en {column_name}')
            return default
        try:
            return int(Decimal(str(value).replace(',', '.')))
        except (InvalidOperation, ValueError, TypeError):
            if job and sheet_result and counter:
                severity = IssueSeverity.ERROR if required else IssueSeverity.WARNING
                if required:
                    counter.errors += 1
                else:
                    counter.warnings += 1
                self._issue(job, sheet_result, severity, sheet_result.sheet_name, row_number, column_name, 'invalid_integer', f'Valor entero invalido en {column_name}', str(value))
            return default

    def _to_decimal(
        self,
        value,
        default=Decimal('0'),
        *,
        job: ImportJob | None = None,
        sheet_result: ImportSheetResult | None = None,
        counter: Counter | None = None,
        row_number: int | None = None,
        column_name: str = '',
        required: bool = False,
    ):
        if value in (None, ''):
            if required and job and sheet_result and counter:
                counter.errors += 1
                self._issue(job, sheet_result, IssueSeverity.ERROR, sheet_result.sheet_name, row_number, column_name, 'missing_required_value', f'Valor requerido ausente en {column_name}')
            return default
        try:
            return Decimal(str(value).replace(' ', '').replace(',', '.'))
        except (InvalidOperation, ValueError, TypeError):
            if job and sheet_result and counter:
                severity = IssueSeverity.ERROR if required else IssueSeverity.WARNING
                if required:
                    counter.errors += 1
                else:
                    counter.warnings += 1
                self._issue(job, sheet_result, severity, sheet_result.sheet_name, row_number, column_name, 'invalid_decimal', f'Valor numerico invalido en {column_name}', str(value))
            return default

    def _to_date(
        self,
        value,
        *,
        job: ImportJob | None = None,
        sheet_result: ImportSheetResult | None = None,
        counter: Counter | None = None,
        row_number: int | None = None,
        column_name: str = '',
        required: bool = False,
    ):
        if value in (None, ''):
            if required and job and sheet_result and counter:
                counter.errors += 1
                self._issue(job, sheet_result, IssueSeverity.ERROR, sheet_result.sheet_name, row_number, column_name, 'missing_required_value', f'Fecha requerida ausente en {column_name}')
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            try:
                converted = from_excel(value)
                return converted.date() if isinstance(converted, datetime) else converted
            except Exception:
                pass
        if isinstance(value, str):
            raw = value.strip()
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
                try:
                    return datetime.strptime(raw, fmt).date()
                except ValueError:
                    continue
        if job and sheet_result and counter:
            severity = IssueSeverity.ERROR if required else IssueSeverity.WARNING
            if required:
                counter.errors += 1
            else:
                counter.warnings += 1
            self._issue(job, sheet_result, severity, sheet_result.sheet_name, row_number, column_name, 'invalid_date', f'Fecha invalida en {column_name}', str(value))
        return None

    def _register_duplicate_key(self, job, sheet_result, counter: Counter, row_number: int, key_name: str, key_value) -> bool:
        normalized_value = self._norm_header(key_value)
        if not normalized_value:
            return False
        sheet_keys = self._seen_keys.setdefault(f'{sheet_result.sheet_name}:{key_name}', {})
        if normalized_value in sheet_keys:
            first_row = sheet_keys[normalized_value]
            counter.warnings += 1
            message = f'Identificador duplicado en el Excel: {key_name}={key_value} ya aparecio en la fila {first_row}.'
            self._issue(job, sheet_result, IssueSeverity.WARNING, sheet_result.sheet_name, row_number, key_name, 'duplicate_in_file', message, str(key_value))
            self._row_result(
                job,
                sheet_result,
                row_number,
                ImportRowAction.WARNING,
                message,
                identifier=str(key_value),
                entity=key_name,
                issue_codes=['duplicate_in_file'],
            )
            return True
        sheet_keys[normalized_value] = row_number
        return False

    def _handle_row_exception(self, job, sheet_result, counter: Counter, row_number: int, exc: Exception, identifier=''):
        logger.exception('Error importando fila %s de hoja %s', row_number, sheet_result.sheet_name)
        counter.errors += 1
        message = f'Fila rechazada por error inesperado: {exc}'
        self._issue(job, sheet_result, IssueSeverity.ERROR, sheet_result.sheet_name, row_number, None, 'row_crash', message)
        self._row_result(job, sheet_result, row_number, ImportRowAction.ERROR, message, identifier=identifier, entity=sheet_result.sheet_name, issue_codes=['row_crash'])

    def _upsert_parcel(self, raw_code: str, counter: Counter, job: ImportJob, sheet_result: ImportSheetResult, row_number: int):
        self._raise_if_cancel_requested(job)
        code = normalize_parcel_code(raw_code)
        if not code:
            counter.errors += 1
            message = 'Fila rechazada: parcela invalida o ausente.'
            self._issue(job, sheet_result, IssueSeverity.ERROR, sheet_result.sheet_name, row_number, 'PARCELA', 'invalid_parcel', message, str(raw_code))
            self._row_result(job, sheet_result, row_number, ImportRowAction.ERROR, message, entity='Parcela', identifier=raw_code, issue_codes=['invalid_parcel'])
            return None

        parcel = Parcel.objects.filter(codigo_parcela_key=code).first()
        if parcel:
            return parcel

        if self.dry_run:
            counter.inserted += 1
            self._row_result(job, sheet_result, row_number, ImportRowAction.CREATED, 'Preview: se crearia la parcela.', entity='Parcela', identifier=code)
            return Parcel(codigo_parcela=code, codigo_parcela_key=code)

        parcel = Parcel.objects.create(codigo_parcela=code)
        counter.inserted += 1
        self._row_result(job, sheet_result, row_number, ImportRowAction.CREATED, 'Parcela creada.', entity='Parcela', identifier=code)
        return parcel

    def _get_or_create_person(
        self,
        *,
        nombre,
        rut='',
        dv='',
        phone1='',
        phone2='',
        email='',
        notes='',
        counter: Counter,
        dry_create=True,
        job: ImportJob | None = None,
        sheet_result: ImportSheetResult | None = None,
        row_number: int | None = None,
    ):
        nombre = (nombre or '').strip()
        rut_norm = normalize_rut_number(rut)
        dv_norm = normalize_rut_dv(dv)
        email_norm = normalize_email(email)

        if rut_norm and dv_norm and not validate_rut(rut_norm, dv_norm):
            if job and sheet_result:
                counter.warnings += 1
                self._issue(job, sheet_result, IssueSeverity.WARNING, sheet_result.sheet_name, row_number, 'RUT', 'invalid_rut', 'RUT invalido; se importara sin digito verificador.', f'{rut}-{dv}')
            dv_norm = ''
        if email_norm:
            try:
                validate_email(email_norm)
            except ValidationError:
                if job and sheet_result:
                    counter.warnings += 1
                    self._issue(job, sheet_result, IssueSeverity.WARNING, sheet_result.sheet_name, row_number, 'EMAIL', 'invalid_email', 'Email invalido; se importara sin correo.', str(email))
                email_norm = ''

        person = None
        if rut_norm:
            person = Person.objects.filter(rut_normalizado=rut_norm).first()
        if not person and email_norm:
            person = Person.objects.filter(email=email_norm, nombre_completo__iexact=nombre).first()
        if not person and nombre:
            person = Person.objects.filter(nombre_completo__iexact=nombre, rut_normalizado='').first()

        defaults = {
            'nombre_completo': nombre or 'Sin nombre',
            'rut': rut_norm,
            'rut_dv': dv_norm,
            'telefono_principal': normalize_phone(phone1),
            'telefono_secundario': normalize_phone(phone2),
            'email': email_norm,
            'notas': notes or '',
            'activo': True,
        }

        if person:
            changes = self._field_diff(person, defaults)
            for field, value in defaults.items():
                if value and getattr(person, field) != value:
                    setattr(person, field, value)
            if changes and not self.dry_run:
                try:
                    person.save()
                    counter.updated += 1
                    if job and sheet_result:
                        self._row_result(
                            job,
                            sheet_result,
                            row_number,
                            ImportRowAction.UPDATED,
                            'Persona actualizada.',
                            entity='Persona',
                            identifier=person.rut_normalizado or person.email or person.nombre_completo,
                            fields_affected=changes,
                        )
                except ValidationError:
                    counter.warnings += 1
            elif changes and self.dry_run and job and sheet_result:
                counter.updated += 1
                self._row_result(
                    job,
                    sheet_result,
                    row_number,
                    ImportRowAction.UPDATED,
                    'Preview: se actualizaria la persona.',
                    entity='Persona',
                    identifier=person.rut_normalizado or person.email or person.nombre_completo,
                    fields_affected=changes,
                )
            return person

        if not dry_create:
            return None

        if self.dry_run:
            counter.inserted += 1
            if job and sheet_result:
                self._row_result(job, sheet_result, row_number, ImportRowAction.CREATED, 'Preview: se crearia la persona.', entity='Persona', identifier=rut_norm or email_norm or nombre)
            return Person(**defaults)

        try:
            person = Person.objects.create(**defaults)
            counter.inserted += 1
            if job and sheet_result:
                self._row_result(job, sheet_result, row_number, ImportRowAction.CREATED, 'Persona creada.', entity='Persona', identifier=person.rut_normalizado or person.email or person.nombre_completo)
            return person
        except ValidationError:
            counter.warnings += 1
            fallback = {**defaults, 'email': '', 'rut_dv': ''}
            try:
                person = Person.objects.create(**fallback)
                counter.inserted += 1
                if job and sheet_result:
                    self._row_result(job, sheet_result, row_number, ImportRowAction.CREATED, 'Persona creada con datos de contacto normalizados.', entity='Persona', identifier=person.rut_normalizado or person.nombre_completo)
                return person
            except ValidationError:
                counter.errors += 1
                if job and sheet_result:
                    self._row_result(job, sheet_result, row_number, ImportRowAction.ERROR, 'Fila rechazada: no se pudo crear la persona.', entity='Persona', identifier=rut_norm or email_norm or nombre, issue_codes=['person_validation_failed'])
                return None

    def _upsert_ownership(self, parcel, person, tipo, counter: Counter, job=None, sheet_result=None, row_number=None):
        if not parcel or not person:
            return
        if self.dry_run and (not getattr(parcel, 'pk', None) or not getattr(person, 'pk', None)):
            counter.inserted += 1
            if job and sheet_result:
                self._row_result(job, sheet_result, row_number, ImportRowAction.CREATED, 'Preview: se crearia la relacion de propietario.', entity='Propiedad', identifier=str(getattr(parcel, 'codigo_parcela', '')))
            return
        lookup = {'parcela': parcel, 'persona': person, 'tipo': tipo}
        existing = ParcelOwnership.objects.filter(**lookup, is_deleted=False).first()
        if existing:
            if not existing.is_active and not self.dry_run:
                existing.is_active = True
                existing.save(update_fields=['is_active', 'updated_at'])
                counter.updated += 1
                if job and sheet_result:
                    self._row_result(job, sheet_result, row_number, ImportRowAction.UPDATED, 'Relacion de propietario reactivada.', entity='Propiedad', identifier=str(parcel))
            else:
                counter.skipped += 1
                if job and sheet_result:
                    self._row_result(job, sheet_result, row_number, ImportRowAction.SKIPPED, 'Relacion de propietario ya existia.', entity='Propiedad', identifier=str(parcel))
            return

        if tipo == OwnershipType.PRINCIPAL:
            active_primary = (
                ParcelOwnership.objects.filter(
                    parcela=parcel,
                    tipo=OwnershipType.PRINCIPAL,
                    is_active=True,
                    is_deleted=False,
                )
                .exclude(persona=person)
                .first()
            )
            if active_primary:
                if self.dry_run:
                    counter.updated += 1
                    if job and sheet_result:
                        self._row_result(
                            job,
                            sheet_result,
                            row_number,
                            ImportRowAction.UPDATED,
                            'Preview: se reemplazaria el propietario principal activo.',
                            entity='Propiedad',
                            identifier=str(parcel),
                        )
                    return
                active_primary.is_active = False
                active_primary.fecha_fin = timezone.now().date()
                active_primary.save(update_fields=['is_active', 'fecha_fin', 'updated_at'])
                counter.updated += 1
                if job and sheet_result:
                    self._row_result(
                        job,
                        sheet_result,
                        row_number,
                        ImportRowAction.UPDATED,
                        'Propietario principal anterior desactivado.',
                        entity='Propiedad',
                        identifier=str(parcel),
                    )
        if self.dry_run:
            counter.inserted += 1
            if job and sheet_result:
                self._row_result(job, sheet_result, row_number, ImportRowAction.CREATED, 'Preview: se crearia la relacion de propietario.', entity='Propiedad', identifier=str(parcel))
            return
        try:
            ParcelOwnership.objects.create(parcela=parcel, persona=person, tipo=tipo, is_active=True)
        except IntegrityError:
            existing_active = ParcelOwnership.objects.filter(parcela=parcel, persona=person, tipo=tipo, is_deleted=False).first()
            if existing_active:
                counter.skipped += 1
                if job and sheet_result:
                    self._row_result(job, sheet_result, row_number, ImportRowAction.SKIPPED, 'Relacion de propietario ya existia.', entity='Propiedad', identifier=str(parcel))
                return
            raise
        counter.inserted += 1
        if job and sheet_result:
            self._row_result(job, sheet_result, row_number, ImportRowAction.CREATED, 'Relacion de propietario creada.', entity='Propiedad', identifier=str(parcel))

    def _parse_datos_propietarios(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'nombre completo', 'rut'])
        if not header_row:
            counter.errors += 1
            self._issue(job, sheet_result, IssueSeverity.ERROR, ws.title, None, None, 'header_not_found', 'No se encontró encabezado en Datos_Propietarios')
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            raw_parcel = ''
            try:
                with transaction.atomic():
                    raw_parcel = self._cell(ws, row, headers, 'parcela')
                    raw_name = self._cell(ws, row, headers, 'nombre completo')
                    if not raw_parcel and not raw_name:
                        continue

                    counter.rows_read += 1
                    if raw_parcel:
                        self._register_duplicate_key(job, sheet_result, counter, row, 'parcela', raw_parcel)
                    parcel = self._upsert_parcel(raw_parcel, counter, job, sheet_result, row)
                    person = self._get_or_create_person(
                        nombre=str(raw_name or '').strip(),
                        rut=self._cell(ws, row, headers, 'rut'),
                        dv=self._cell(ws, row, headers, 'dv'),
                        phone1=self._cell(ws, row, headers, 'telefono fijo'),
                        phone2=self._cell(ws, row, headers, 'telefono movil'),
                        email=self._cell(ws, row, headers, 'e mail', 'email'),
                        notes=self._cell(ws, row, headers, 'obs esp'),
                        counter=counter,
                        job=job,
                        sheet_result=sheet_result,
                        row_number=row,
                    )
                    if parcel and person:
                        self._upsert_ownership(parcel, person, OwnershipType.PRINCIPAL, counter, job=job, sheet_result=sheet_result, row_number=row)
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=raw_parcel if 'raw_parcel' in locals() else '')

    def _parse_otros_duenos(self, ws, job, sheet_result, counter: Counter):
        for row in self._iter_data_row_numbers(ws, 2, max_col=19):
            parcela = ''
            try:
                with transaction.atomic():
                    parcela = ws.cell(row=row, column=1).value
                    if not parcela:
                        continue
                    counter.rows_read += 1
                    self._register_duplicate_key(job, sheet_result, counter, row, 'parcela', parcela)
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)

                    for offset in (2, 5, 8, 11, 14, 17):
                        rut = ws.cell(row=row, column=offset).value
                        dv = ws.cell(row=row, column=offset + 1).value
                        nombre = ws.cell(row=row, column=offset + 2).value
                        if not nombre and not rut:
                            continue
                        person = self._get_or_create_person(
                            nombre=str(nombre or '').strip(),
                            rut=rut,
                            dv=dv,
                            counter=counter,
                            dry_create=True,
                            job=job,
                            sheet_result=sheet_result,
                            row_number=row,
                        )
                        if parcel and person:
                            self._upsert_ownership(parcel, person, OwnershipType.COPROPIETARIO, counter, job=job, sheet_result=sheet_result, row_number=row)
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=parcela)

    def _parse_residentes(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'residente'])
        if not header_row:
            counter.warnings += 1
            self._issue(job, sheet_result, IssueSeverity.WARNING, ws.title, None, None, 'header_not_found', 'No se detectó encabezado en RESIDENTES')
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            parcela = ''
            try:
                with transaction.atomic():
                    parcela = self._cell(ws, row, headers, 'parcela')
                    estado_residente = self._cell(ws, row, headers, 'residente')
                    observaciones = self._cell(ws, row, headers, 'observaciones')
                    if not parcela:
                        continue

                    counter.rows_read += 1
                    self._register_duplicate_key(job, sheet_result, counter, row, 'parcela', parcela)
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
                    if not parcel:
                        continue

                    tipo = ResidentType.CUIDADOR if 'cuid' in str(estado_residente or '').lower() else ResidentType.RESIDENTE
                    active = 'INACT' not in str(estado_residente or '').upper()
                    obs = str(observaciones or '').strip()
                    if self.dry_run and not getattr(parcel, 'pk', None):
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia el residente.', entity='Residente', identifier=str(parcela))
                        continue
                    existing = ParcelResident.objects.filter(
                        parcela=parcel,
                        persona__isnull=True,
                        tipo_residencia=tipo,
                        observaciones=obs,
                        is_deleted=False,
                    ).first()
                    if existing:
                        if existing.is_active != active and not self.dry_run:
                            existing.is_active = active
                            existing.save(update_fields=['is_active', 'updated_at'])
                            counter.updated += 1
                            self._row_result(job, sheet_result, row, ImportRowAction.UPDATED, 'Residente actualizado.', entity='Residente', identifier=str(parcel), fields_affected=[{'field': 'is_active', 'before': str(not active), 'after': str(active)}])
                        else:
                            counter.skipped += 1
                            self._row_result(job, sheet_result, row, ImportRowAction.SKIPPED, 'Residente ya existia sin cambios.', entity='Residente', identifier=str(parcel))
                        continue
                    if self.dry_run:
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia el residente.', entity='Residente', identifier=str(parcel))
                        continue
                    ParcelResident.objects.create(
                        parcela=parcel,
                        persona=None,
                        tipo_residencia=tipo,
                        is_active=active,
                        observaciones=obs,
                    )
                    counter.inserted += 1
                    self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Residente creado.', entity='Residente', identifier=str(parcel))
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=parcela)

    def _parse_vehiculos(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'ppu'])
        if not header_row:
            counter.errors += 1
            self._issue(job, sheet_result, IssueSeverity.ERROR, ws.title, None, None, 'header_not_found', 'No se detectó encabezado en PPU_LOGOS')
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            parcela = ''
            ppu = ''
            try:
                with transaction.atomic():
                    parcela = self._cell(ws, row, headers, 'parcela')
                    ppu = self._cell(ws, row, headers, 'ppu')
                    if not parcela and not ppu:
                        continue
                    if not ppu:
                        counter.rows_read += 1
                        counter.errors += 1
                        message = 'Fila rechazada: PPU/patente ausente.'
                        self._issue(job, sheet_result, IssueSeverity.ERROR, ws.title, row, 'PPU', 'missing_ppu', message)
                        self._row_result(job, sheet_result, row, ImportRowAction.ERROR, message, entity='Vehiculo', identifier=str(parcela), issue_codes=['missing_ppu'])
                        continue
                    counter.rows_read += 1
                    self._register_duplicate_key(job, sheet_result, counter, row, 'parcela_ppu', f'{parcela}:{ppu}')
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
                    if not parcel:
                        continue

                    ppu_norm = ''.join(ch for ch in str(ppu).upper() if ch.isalnum())
                    defaults = {
                        'marca': str(self._cell(ws, row, headers, 'marca') or '').strip(),
                        'tipo': str(self._cell(ws, row, headers, 'tipo') or '').strip(),
                        'color': str(self._cell(ws, row, headers, 'color') or '').strip(),
                        'codigo_acceso': str(self._cell(ws, row, headers, 'codigo') or '').strip(),
                        'ppu': str(ppu).strip().upper(),
                        'activo': True,
                    }
                    if self.dry_run and not getattr(parcel, 'pk', None):
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia el vehiculo.', entity='Vehiculo', identifier=ppu_norm)
                        continue
                    existing = Vehicle.objects.filter(parcela=parcel, ppu_normalizado=ppu_norm, is_deleted=False).first()
                    if existing:
                        changes = self._field_diff(existing, defaults)
                        if changes and not self.dry_run:
                            for key, val in defaults.items():
                                if val:
                                    setattr(existing, key, val)
                            existing.save()
                            counter.updated += 1
                            self._row_result(job, sheet_result, row, ImportRowAction.UPDATED, 'Vehiculo actualizado.', entity='Vehiculo', identifier=ppu_norm, fields_affected=changes)
                        elif changes and self.dry_run:
                            counter.updated += 1
                            self._row_result(job, sheet_result, row, ImportRowAction.UPDATED, 'Preview: se actualizaria el vehiculo.', entity='Vehiculo', identifier=ppu_norm, fields_affected=changes)
                        else:
                            counter.skipped += 1
                            self._row_result(job, sheet_result, row, ImportRowAction.SKIPPED, 'Vehiculo ya existia sin cambios.', entity='Vehiculo', identifier=ppu_norm)
                        continue

                    if self.dry_run:
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia el vehiculo.', entity='Vehiculo', identifier=ppu_norm)
                        continue
                    Vehicle.objects.create(parcela=parcel, **defaults)
                    counter.inserted += 1
                    self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Vehiculo creado.', entity='Vehiculo', identifier=ppu_norm)
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=f'{parcela}:{ppu}')

    def _parse_mora_gc(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'mora cg uf', 'total pesos'])
        if not header_row:
            counter.warnings += 1
            self._issue(job, sheet_result, IssueSeverity.WARNING, ws.title, None, None, 'header_not_found', 'No se detectó encabezado en Mora GC')
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            parcela = ''
            try:
                with transaction.atomic():
                    parcela = self._cell(ws, row, headers, 'parcela')
                    if not parcela:
                        continue
                    counter.rows_read += 1
                    self._register_duplicate_key(job, sheet_result, counter, row, 'parcela', parcela)
                    total_pesos = self._to_decimal(self._cell(ws, row, headers, 'total pesos'), default=None, job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='TOTAL PESOS', required=True)
                    if total_pesos is None:
                        self._row_result(job, sheet_result, row, ImportRowAction.ERROR, 'Fila rechazada: total pesos invalido.', entity='Mora GC', identifier=str(parcela), issue_codes=['invalid_decimal'])
                        continue
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
                    if not parcel:
                        continue

                    defaults = {
                        'numero_gastos_comunes': self._to_int(self._cell(ws, row, headers, 'n gastos comunes'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='N GASTOS COMUNES'),
                        'mora_uf': self._to_decimal(self._cell(ws, row, headers, 'mora cg uf'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='MORA CG UF'),
                        'interes_mora_uf': self._to_decimal(self._cell(ws, row, headers, 'interes mora uf'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='INTERES MORA UF'),
                        'total_uf': self._to_decimal(self._cell(ws, row, headers, 'total uf'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='TOTAL UF'),
                        'total_pesos': total_pesos,
                        'estado_pago': 'PENDIENTE',
                    }
                    if self.dry_run and not getattr(parcel, 'pk', None):
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia deuda de gastos comunes.', entity='Mora GC', identifier=str(parcela))
                        continue
                    duplicate = CommonExpenseDebt.objects.filter(
                        parcela=parcel,
                        numero_gastos_comunes=defaults['numero_gastos_comunes'],
                        total_pesos=defaults['total_pesos'],
                        total_uf=defaults['total_uf'],
                        is_deleted=False,
                    ).exists()
                    if duplicate:
                        counter.skipped += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.SKIPPED, 'Deuda GC duplicada; se omitio.', entity='Mora GC', identifier=str(parcela))
                        continue
                    if self.dry_run:
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia deuda de gastos comunes.', entity='Mora GC', identifier=str(parcela))
                        continue
                    CommonExpenseDebt.objects.create(parcela=parcel, **defaults)
                    counter.inserted += 1
                    self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Deuda GC creada.', entity='Mora GC', identifier=str(parcela))
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=parcela)

    def _parse_deudas_ays(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'total deuda'])
        if not header_row:
            counter.warnings += 1
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            parcela = ''
            try:
                with transaction.atomic():
                    parcela = self._cell(ws, row, headers, 'parcela')
                    if not parcela:
                        continue
                    counter.rows_read += 1
                    self._register_duplicate_key(job, sheet_result, counter, row, 'parcela', parcela)
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
                    if not parcel:
                        continue

                    saldo_total = self._to_decimal(self._cell(ws, row, headers, 'total deuda'), default=None, job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='TOTAL DEUDA', required=True)
                    if saldo_total is None:
                        self._row_result(job, sheet_result, row, ImportRowAction.ERROR, 'Fila rechazada: total deuda invalido.', entity='Deuda AYS', identifier=str(parcela), issue_codes=['invalid_decimal'])
                        continue
                    defaults = {
                        'tipo_servicio': ServiceType.AYS,
                        'numero_boletas': self._to_int(self._cell(ws, row, headers, 'boletas'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='BOLETAS'),
                        'monto_total': self._to_decimal(self._cell(ws, row, headers, 'a s total', 'total'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='TOTAL'),
                        'convenios': self._to_decimal(self._cell(ws, row, headers, 'convenios'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='CONVENIOS'),
                        'anticipos': self._to_decimal(self._cell(ws, row, headers, 'anticipos'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='ANTICIPOS'),
                        'saldo_total': saldo_total,
                        'estado_pago': 'PENDIENTE' if saldo_total > 0 else 'PAGADO',
                        'observaciones': str(self._cell(ws, row, headers, 'comentarios') or ''),
                    }
                    if self.dry_run and not getattr(parcel, 'pk', None):
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia deuda AYS.', entity='Deuda AYS', identifier=str(parcela))
                        continue
                    existing = ServiceDebt.objects.filter(
                        parcela=parcel,
                        tipo_servicio=ServiceType.AYS,
                        is_deleted=False,
                    ).order_by('-created_at').first()
                    if existing and existing.saldo_total == defaults['saldo_total'] and existing.numero_boletas == defaults['numero_boletas']:
                        counter.skipped += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.SKIPPED, 'Deuda AYS sin cambios; se omitio.', entity='Deuda AYS', identifier=str(parcela))
                        continue
                    if self.dry_run:
                        action = ImportRowAction.UPDATED if existing else ImportRowAction.CREATED
                        counter.updated += 1 if existing else 0
                        counter.inserted += 0 if existing else 1
                        self._row_result(job, sheet_result, row, action, 'Preview: se actualizaria deuda AYS.' if existing else 'Preview: se crearia deuda AYS.', entity='Deuda AYS', identifier=str(parcela))
                        continue
                    if existing:
                        changes = self._field_diff(existing, defaults)
                        for key, value in defaults.items():
                            setattr(existing, key, value)
                        existing.save()
                        counter.updated += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.UPDATED, 'Deuda AYS actualizada.', entity='Deuda AYS', identifier=str(parcela), fields_affected=changes)
                    else:
                        ServiceDebt.objects.create(parcela=parcel, **defaults)
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Deuda AYS creada.', entity='Deuda AYS', identifier=str(parcela))
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=parcela)

    def _parse_mora_convenio(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'total mora'])
        if not header_row:
            counter.warnings += 1
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            parcela = ''
            try:
                with transaction.atomic():
                    parcela = self._cell(ws, row, headers, 'parcela')
                    if not parcela:
                        continue
                    counter.rows_read += 1
                    self._register_duplicate_key(job, sheet_result, counter, row, 'parcela', parcela)
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
                    if not parcel:
                        continue

                    saldo = self._to_decimal(self._cell(ws, row, headers, 'total mora'), default=None, job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='TOTAL MORA', required=True)
                    if saldo is None:
                        self._row_result(job, sheet_result, row, ImportRowAction.ERROR, 'Fila rechazada: total mora invalido.', entity='Mora Convenio', identifier=str(parcela), issue_codes=['invalid_decimal'])
                        continue
                    if saldo <= 0:
                        counter.skipped += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.SKIPPED, 'Saldo de mora no positivo; se omitio.', entity='Mora Convenio', identifier=str(parcela))
                        continue

                    defaults = {
                        'empresa': str(self._cell(ws, row, headers, 'cobranza') or ''),
                        'tipo': 'MORA_CONVENIO',
                        'detalle': f"GC: {self._to_int(self._cell(ws, row, headers, 'n gc'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='N GC')}",
                        'saldo_monto': saldo,
                        'estado_pago': 'PENDIENTE',
                    }
                    if self.dry_run and not getattr(parcel, 'pk', None):
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia convenio en mora.', entity='Mora Convenio', identifier=str(parcela))
                        continue
                    duplicate = PaymentAgreement.objects.filter(
                        parcela=parcel,
                        empresa=defaults['empresa'],
                        tipo=defaults['tipo'],
                        detalle=defaults['detalle'],
                        saldo_monto=defaults['saldo_monto'],
                        is_deleted=False,
                    ).exists()
                    if duplicate:
                        counter.skipped += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.SKIPPED, 'Convenio en mora duplicado; se omitio.', entity='Mora Convenio', identifier=str(parcela))
                        continue
                    if self.dry_run:
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia convenio en mora.', entity='Mora Convenio', identifier=str(parcela))
                        continue
                    PaymentAgreement.objects.create(parcela=parcel, **defaults)
                    counter.inserted += 1
                    self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Convenio en mora creado.', entity='Mora Convenio', identifier=str(parcela))
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=parcela)

    def _parse_multas(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'empresa', 'saldo monto'])
        if not header_row:
            counter.warnings += 1
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            parcela = ''
            try:
                with transaction.atomic():
                    parcela = self._cell(ws, row, headers, 'parcela')
                    if not parcela:
                        continue
                    counter.rows_read += 1
                    self._register_duplicate_key(job, sheet_result, counter, row, 'parcela', parcela)
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
                    if not parcel:
                        continue

                    saldo = self._to_decimal(self._cell(ws, row, headers, 'saldo monto'), default=None, job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='SALDO MONTO', required=True)
                    if saldo is None:
                        self._row_result(job, sheet_result, row, ImportRowAction.ERROR, 'Fila rechazada: saldo monto invalido.', entity='Multa', identifier=str(parcela), issue_codes=['invalid_decimal'])
                        continue
                    defaults = {
                        'empresa': str(self._cell(ws, row, headers, 'empresa') or ''),
                        'tipo': str(self._cell(ws, row, headers, 'tipo') or ''),
                        'fecha_emision': self._to_date(self._cell(ws, row, headers, 'emision'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='EMISION'),
                        'fecha_vencimiento': self._to_date(self._cell(ws, row, headers, 'vencimiento'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='VENCIMIENTO'),
                        'detalle': str(self._cell(ws, row, headers, 'detalle') or ''),
                        'saldo_monto': saldo,
                        'estado_pago': 'PENDIENTE' if saldo > 0 else 'PAGADO',
                    }
                    if self.dry_run and not getattr(parcel, 'pk', None):
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia multa impaga.', entity='Multa', identifier=str(parcela))
                        continue
                    duplicate = UnpaidFine.objects.filter(
                        parcela=parcel,
                        empresa=defaults['empresa'],
                        tipo=defaults['tipo'],
                        fecha_vencimiento=defaults['fecha_vencimiento'],
                        saldo_monto=defaults['saldo_monto'],
                        is_deleted=False,
                    ).exists()
                    if duplicate:
                        counter.skipped += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.SKIPPED, 'Multa duplicada; se omitio.', entity='Multa', identifier=str(parcela))
                        continue
                    if self.dry_run:
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia multa impaga.', entity='Multa', identifier=str(parcela))
                        continue
                    UnpaidFine.objects.create(parcela=parcel, **defaults)
                    counter.inserted += 1
                    self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Multa impaga creada.', entity='Multa', identifier=str(parcela))
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=parcela)

    def _parse_cortes(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['cliente', 'estado'])
        if not header_row:
            counter.warnings += 1
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            parcela = ''
            try:
                with transaction.atomic():
                    parcela = self._cell(ws, row, headers, 'cliente')
                    if not parcela:
                        continue
                    counter.rows_read += 1
                    self._register_duplicate_key(job, sheet_result, counter, row, 'cliente', parcela)
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
                    if not parcel:
                        continue

                    estado = str(self._cell(ws, row, headers, 'estado') or '').strip()
                    te1 = str(self._cell(ws, row, headers, 'te1 vencimiento') or '').strip()
                    corte_luz = str(self._cell(ws, row, headers, 'corte luz') or '').strip()
                    corte_ap = str(self._cell(ws, row, headers, 'corte ap') or '').strip()

                    tipo = CutType.AYS
                    if corte_luz and not corte_ap:
                        tipo = CutType.LUZ
                    elif corte_ap and not corte_luz:
                        tipo = CutType.AGUA

                    defaults = {
                        'tipo_corte': tipo,
                        'estado': estado,
                        'motivo': te1,
                        'fecha': self._to_date(self._cell(ws, row, headers, 'fecha'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='FECHA'),
                        'activo': True,
                    }
                    if self.dry_run and not getattr(parcel, 'pk', None):
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia corte vigente.', entity='Corte', identifier=str(parcela))
                        continue
                    existing = ServiceCut.objects.filter(
                        parcela=parcel,
                        tipo_corte=defaults['tipo_corte'],
                        fecha=defaults['fecha'],
                        motivo=defaults['motivo'],
                        is_deleted=False,
                    ).first()
                    if existing:
                        changed = existing.estado != defaults['estado'] or existing.activo != defaults['activo']
                        if changed and not self.dry_run:
                            changes = self._field_diff(existing, defaults)
                            existing.estado = defaults['estado']
                            existing.activo = defaults['activo']
                            existing.save(update_fields=['estado', 'activo', 'updated_at'])
                            counter.updated += 1
                            self._row_result(job, sheet_result, row, ImportRowAction.UPDATED, 'Corte actualizado.', entity='Corte', identifier=str(parcela), fields_affected=changes)
                        elif changed and self.dry_run:
                            counter.updated += 1
                            self._row_result(job, sheet_result, row, ImportRowAction.UPDATED, 'Preview: se actualizaria corte vigente.', entity='Corte', identifier=str(parcela))
                        else:
                            counter.skipped += 1
                            self._row_result(job, sheet_result, row, ImportRowAction.SKIPPED, 'Corte ya existia sin cambios.', entity='Corte', identifier=str(parcela))
                        continue
                    if self.dry_run:
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia corte vigente.', entity='Corte', identifier=str(parcela))
                        continue
                    ServiceCut.objects.create(parcela=parcel, **defaults)
                    counter.inserted += 1
                    self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Corte vigente creado.', entity='Corte', identifier=str(parcela))
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=parcela)

    def _parse_historico_ays(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'solicitante', 'descripcion'])
        if not header_row:
            counter.warnings += 1
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            parcela = ''
            try:
                with transaction.atomic():
                    parcela = self._cell(ws, row, headers, 'parcela')
                    if not parcela:
                        continue
                    counter.rows_read += 1
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
                    if not parcel:
                        continue

                    defaults = {
                        'numero_orden': str(self._cell(ws, row, headers, 'orden') or ''),
                        'solicitante': str(self._cell(ws, row, headers, 'solicitante') or ''),
                        'resultado': str(self._cell(ws, row, headers, 'realizado') or ''),
                        'descripcion': str(self._cell(ws, row, headers, 'descripcion') or ''),
                        'fecha_ingreso': self._to_date(self._cell(ws, row, headers, 'fecha ingreso'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='FECHA INGRESO'),
                        'fecha_ejecucion': self._to_date(self._cell(ws, row, headers, 'fecha ejecucion'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='FECHA EJECUCION'),
                        'ejecutante': str(self._cell(ws, row, headers, 'ejecutante') or ''),
                        'lugar_corte_reposicion': str(self._cell(ws, row, headers, 'lugar de corte') or ''),
                        'observaciones': str(self._cell(ws, row, headers, 'obvervaciones', 'observaciones') or ''),
                    }
                    identifier = defaults['numero_orden'] or str(parcela)
                    self._register_duplicate_key(job, sheet_result, counter, row, 'orden_parcela', f"{parcela}:{identifier}:{defaults['descripcion']}")
                    if self.dry_run and not getattr(parcel, 'pk', None):
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia historico AYS.', entity='Historico AYS', identifier=identifier)
                        continue
                    duplicate = ServiceHistory.objects.filter(
                        parcela=parcel,
                        numero_orden=defaults['numero_orden'],
                        descripcion=defaults['descripcion'],
                        fecha_ingreso=defaults['fecha_ingreso'],
                        is_deleted=False,
                    ).exists()
                    if duplicate:
                        counter.skipped += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.SKIPPED, 'Historico AYS duplicado; se omitio.', entity='Historico AYS', identifier=identifier)
                        continue
                    if self.dry_run:
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia historico AYS.', entity='Historico AYS', identifier=identifier)
                        continue
                    ServiceHistory.objects.create(parcela=parcel, **defaults)
                    counter.inserted += 1
                    self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Historico AYS creado.', entity='Historico AYS', identifier=identifier)
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=parcela)

    def _parse_anotaciones(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'fecha', 'anotacion'])
        if not header_row:
            counter.warnings += 1
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            parcela = ''
            try:
                with transaction.atomic():
                    parcela = self._cell(ws, row, headers, 'parcela')
                    texto = self._cell(ws, row, headers, 'anotacion')
                    if not parcela and not texto:
                        continue
                    if not texto:
                        counter.rows_read += 1
                        counter.errors += 1
                        message = 'Fila rechazada: anotacion ausente.'
                        self._issue(job, sheet_result, IssueSeverity.ERROR, ws.title, row, 'ANOTACION', 'missing_note', message)
                        self._row_result(job, sheet_result, row, ImportRowAction.ERROR, message, entity='Anotacion', identifier=str(parcela), issue_codes=['missing_note'])
                        continue

                    counter.rows_read += 1
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
                    if not parcel:
                        continue
                    event_date = self._to_date(self._cell(ws, row, headers, 'fecha'), job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='FECHA')
                    normalized_text = str(texto).strip()
                    identifier = f'{parcela}:{normalized_text[:40]}'
                    self._register_duplicate_key(job, sheet_result, counter, row, 'anotacion', identifier)
                    if self.dry_run and not getattr(parcel, 'pk', None):
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia anotacion.', entity='Anotacion', identifier=identifier)
                        continue
                    duplicate = AdministrativeNote.objects.filter(
                        parcela=parcel,
                        texto=normalized_text,
                        fecha_evento=event_date,
                        is_deleted=False,
                    ).exists()
                    if duplicate:
                        counter.skipped += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.SKIPPED, 'Anotacion duplicada; se omitio.', entity='Anotacion', identifier=identifier)
                        continue
                    if self.dry_run:
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Preview: se crearia anotacion.', entity='Anotacion', identifier=identifier)
                        continue

                    AdministrativeNote.objects.create(
                        parcela=parcel,
                        tipo=NoteType.ADMINISTRATIVA,
                        texto=normalized_text,
                        fecha_evento=event_date,
                    )
                    counter.inserted += 1
                    self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Anotacion creada.', entity='Anotacion', identifier=identifier)
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=parcela)

    def _parse_obras(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela n', 'cortafuego', 'limpieza'])
        if not header_row:
            counter.warnings += 1
            return

        for row in self._iter_data_row_numbers(ws, header_row + 1, max_col=max(headers.values())):
            parcela = ''
            try:
                with transaction.atomic():
                    parcela = self._cell(ws, row, headers, 'parcela n')
                    if not parcela:
                        continue

                    counter.rows_read += 1
                    self._register_duplicate_key(job, sheet_result, counter, row, 'parcela', parcela)
                    parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
                    if not parcel:
                        continue

                    defaults = {
                        'deshabitada': str(self._cell(ws, row, headers, 'deshabitada') or ''),
                        'cercada': str(self._cell(ws, row, headers, 'cercada') or ''),
                        'sucia': str(self._cell(ws, row, headers, 'sucia') or ''),
                        'casas': str(self._cell(ws, row, headers, 'casas') or ''),
                        'otra_construccion': str(self._cell(ws, row, headers, 'otra const') or ''),
                        'cumplen': str(self._cell(ws, row, headers, 'cumplen') or ''),
                        'cortafuego': str(self._cell(ws, row, headers, 'cortafuego') or ''),
                        'limpieza': str(self._cell(ws, row, headers, 'limpieza') or ''),
                        'foco_incendio': str(self._cell(ws, row, headers, 'foco incend') or ''),
                        'atributo_kpi': self._to_decimal(self._cell(ws, row, headers, 'atributo kpi'), default=None, job=job, sheet_result=sheet_result, counter=counter, row_number=row, column_name='ATRIBUTO KPI'),
                        'kpi': str(self._cell(ws, row, headers, 'kpi') or ''),
                        'estado_actual': str(self._cell(ws, row, headers, 'estado actual') or ''),
                        'rol_sii': str(self._cell(ws, row, headers, 'rol') or ''),
                        'certificado_obras': str(self._cell(ws, row, headers, 'certificado obras') or ''),
                        'permiso_dom': str(self._cell(ws, row, headers, 'permiso dom') or ''),
                    }

                    if self.dry_run:
                        action = ImportRowAction.UPDATED if getattr(parcel, 'pk', None) and ParcelWorkStatus.objects.filter(parcela=parcel).exists() else ImportRowAction.CREATED
                        if action == ImportRowAction.UPDATED:
                            counter.updated += 1
                        else:
                            counter.inserted += 1
                        self._row_result(job, sheet_result, row, action, 'Preview: se sincronizaria estado de obras.', entity='Obras', identifier=str(parcela))
                        continue

                    _, created = ParcelWorkStatus.objects.update_or_create(parcela=parcel, defaults=defaults)
                    if created:
                        counter.inserted += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.CREATED, 'Estado de obras creado.', entity='Obras', identifier=str(parcela))
                    else:
                        counter.updated += 1
                        self._row_result(job, sheet_result, row, ImportRowAction.UPDATED, 'Estado de obras actualizado.', entity='Obras', identifier=str(parcela), fields_affected=list(defaults.keys()))
            except (ValidationError, IntegrityError, ValueError) as exc:
                self._handle_row_exception(job, sheet_result, counter, row, exc, identifier=parcela)

